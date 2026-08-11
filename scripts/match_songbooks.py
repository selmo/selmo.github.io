#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""_data/scores.yml 의 각 곡에 성가집 번호를 매칭.

- 야훼이레 3판: resources/야훼이레-수록곡목록-3판.xlsx (총괄 시트 + Sheet1 색인)
- 하늘바다 (가톨릭 어린이 찬양집): resources/하늘바다.md
  · 미사곡: 파트별 '번호 | 작곡/편곡' 표 -> (파트, 작곡가)로 매칭
  · 일반곡: '제목별 차례' 표 -> 곡명으로 매칭

매칭 결과를 scores.yml 에 yahure / haneulbada 필드로 기록.
곡명이 같아도 작곡가가 다르면 다른 곡이므로, 작곡가 정보가 있을 때만
동명이곡을 확정한다. 확정 불가하면 비워 둔다(추측 금지).
"""
import openpyxl, yaml, re, unicodedata, sys
from collections import defaultdict

XLSX = "resources/야훼이레-수록곡목록-3판.xlsx"
HB_MD = "resources/하늘바다.md"
SCORES = "_data/scores.yml"


def norm(s):
    """곡명 비교용 정규화: 공백/문장부호 제거 + NFC."""
    if not s:
        return ""
    s = unicodedata.normalize("NFC", str(s))
    return re.sub(r"[\s·・,\.\'\"’‘“”\(\)\[\]!?~\-—]", "", s).lower()


def norm_composer(s):
    """작곡가 비교용: 세례명 제거하고 성명만 (예: '최태형 안셀모' -> '최태형')."""
    if not s:
        return ""
    s = unicodedata.normalize("NFC", str(s))
    m = re.match(r"^([가-힣]{2,4})\s", s)
    return m.group(1) if m else re.sub(r"\s", "", s)


# ---------- 야훼이레 ----------
def load_yahure():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["총괄"]
    # (번호, 곡명, 작곡자)
    entries = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if isinstance(r[1], int) and r[3]:
            entries.append((r[1], r[3], r[5]))
    # Sheet1 색인(곡명->번호) 보강: 총괄에 없는 곡명만
    have = {norm(e[1]) for e in entries}
    ws1 = wb["Sheet1"]
    for r in ws1.iter_rows(values_only=True):
        if r[0] and isinstance(r[1], int) and norm(r[0]) not in have:
            entries.append((r[1], r[0], None))
    by_title = defaultdict(list)
    for num, title, comp in entries:
        by_title[norm(title)].append((num, title, comp))
    return by_title


# ---------- 하늘바다 ----------
# 하늘바다 미사 파트 헤딩 -> scores.yml mass_part
HB_PART_MAP = {
    "자비송": "kyrie",
    "대영광송": "gloria",
    "화답송": "responsorial_psalm",
    "복음환호송": "gospel_acclamation",
    "보편지향기도": "prayer_of_faithful",
    "거룩하시도다": "sanctus",
    "신앙의 신비여": "mystery",
    "아멘": "amen",
    "주님의 기도": "lords_prayer",
    "주님께 나라와": "doxology",
    "하느님의 어린양": "agnus",
    "영광송": "doxology",
}


def load_haneulbada():
    text = open(HB_MD, encoding="utf-8").read()
    lines = text.split("\n")

    # 미사곡: 파트별 (번호, 작곡가)
    mass = defaultdict(list)          # mass_part -> [(번호, 작곡가원문)]
    # 일반곡: 곡명 -> 번호
    titles = {}

    section = None       # 〈미사곡〉 / 〈제목별 차례〉 등
    part = None
    for ln in lines:
        h2 = re.match(r"^##\s+〈(.+?)〉", ln)
        if h2:
            section = h2.group(1)
            part = None
            continue
        h3 = re.match(r"^###\s+○\s+(.+?)\s*$", ln)
        if h3:
            part = h3.group(1).strip()
            continue
        row = re.match(r"^\|\s*(.+?)\s*\|\s*(.+?)\s*\|\s*$", ln)
        row4 = re.match(r"^\|\s*(.+?)\s*\|\s*(.+?)\s*\|\s*(.+?)\s*\|\s*(.+?)\s*\|\s*$", ln)

        if section == "미사곡" and part and row and not row4:
            num, comp = row.group(1), row.group(2)
            if num.isdigit():
                mp = HB_PART_MAP.get(part)
                if mp:
                    mass[mp].append((int(num), comp))
        elif section and "제목별" in section and row4:
            # | 번호 | 곡명 | 번호 | 곡명 |
            a_num, a_t, b_num, b_t = row4.groups()
            for n, t in ((a_num, a_t), (b_num, b_t)):
                if n.strip().isdigit():
                    titles.setdefault(norm(t), (int(n), t.strip()))
    return mass, titles


def main():
    yh = load_yahure()
    hb_mass, hb_titles = load_haneulbada()

    print(f"야훼이레 색인: {len(yh)}개 곡명")
    print(f"하늘바다 미사곡: {sum(len(v) for v in hb_mass.values())}개, 일반곡: {len(hb_titles)}개")
    print()

    d = yaml.safe_load(open(SCORES, encoding="utf-8"))
    items = d["items"]

    # group 단위로 매칭 (같은 곡 = 같은 번호)
    groups = defaultdict(list)
    for it in items:
        groups[it["group"]].append(it)

    report = []
    for g, gi in groups.items():
        first = gi[0]
        title, comp, mp = first["title"], first["composer"], first["mass_part"]
        t = norm(title)

        # --- 야훼이레 ---
        yh_num = first.get("yahure")   # 파일명에서 이미 추출된 값 우선
        yh_src = "파일명" if yh_num else ""
        if not yh_num:
            cands = yh.get(t, [])
            if comp:
                for num, nm, cp in cands:
                    if norm_composer(cp) == norm_composer(comp):
                        yh_num, yh_src = num, "곡명+작곡가"
                        break
            if not yh_num and len(cands) == 1 and not comp:
                # 동명이곡이 없으면 곡명만으로 확정
                yh_num, yh_src = cands[0][0], "곡명"

        # --- 하늘바다 ---
        hb_num = None
        hb_src = ""
        if mp and mp in hb_mass:
            # 미사곡: 파트 안에서 작곡가 일치
            if comp:
                for num, cp in hb_mass[mp]:
                    # '신지은, 이지혜' / '이종철(하느님 말씀…)' 형태 처리
                    cp_names = re.split(r"[,()]", cp)
                    cp_names = [norm_composer(x.strip()) for x in cp_names if x.strip()]
                    for cn in [norm_composer(x) for x in re.split(r"[·,]", comp)]:
                        if cn and cn in cp_names:
                            hb_num, hb_src = num, "미사파트+작곡가"
                            break
                    if hb_num:
                        break
        if not hb_num and t in hb_titles:
            hb_num, hb_src = hb_titles[t][0], "곡명"

        for it in gi:
            it["yahure"] = yh_num
            it["haneulbada"] = hb_num
        report.append((title, comp, mp, yh_num, yh_src, hb_num, hb_src))

    # 리포트
    print("=== 매칭 결과 ===")
    print(f"{'곡명':<28} {'작곡':<12} {'야훼이레':>8} {'하늘바다':>8}")
    for title, comp, mp, yn, ys, hn, hs in report:
        yt = f"{yn}({ys})" if yn else "-"
        ht = f"{hn}({hs})" if hn else "-"
        print(f"  {title:<28} {comp or '-':<12} {yt:>14} {ht:>18}")

    yh_hit = sum(1 for r in report if r[3])
    hb_hit = sum(1 for r in report if r[5])
    print(f"\n곡 {len(report)}개 중 — 야훼이레 {yh_hit} · 하늘바다 {hb_hit}")

    if "--write" in sys.argv:
        write_yaml(items)
        print("\n_data/scores.yml 갱신 완료")


def esc(v):
    if v is None:
        return "null"
    if isinstance(v, int):
        return str(v)
    s = str(v).replace("\\", "\\\\").replace('"', '\\"')
    return f'"{s}"'


def write_yaml(items):
    with open(SCORES, "w", encoding="utf-8") as out:
        out.write("# OnE-Scores PDF 메타데이터\n")
        out.write("# path = 파일 시스템 원본(NFD 포함 그대로). url = 퍼센트 인코딩 링크.\n")
        out.write("# group 이 같으면 같은 곡의 변형(Score/ChordChart/Chorus/키) -> 한 줄에 여러 링크.\n")
        out.write("# yahure = 야훼이레 3판 번호, haneulbada = 하늘바다(어린이 찬양집) 번호.\n")
        out.write("items:\n")
        for it in items:
            for k in ("path", "url", "title", "composer"):
                out.write(f'  {"- " if k == "path" else "  "}{k}: {esc(it[k])}\n')
            out.write(f'    category: {it["category"]}\n')
            out.write(f'    mass_part: {esc(it["mass_part"])}\n')
            out.write(f'    type: {it["type"]}\n')
            out.write(f'    yahure: {esc(it["yahure"])}\n')
            out.write(f'    haneulbada: {esc(it.get("haneulbada"))}\n')
            out.write(f'    key: {esc(it["key"])}\n')
            out.write(f'    note: {esc(it["note"])}\n')
            out.write(f'    group: {esc(it["group"])}\n')
            out.write(f'    label: {esc(it["label"])}\n')


if __name__ == "__main__":
    main()
